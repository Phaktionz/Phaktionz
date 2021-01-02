from openpyxl import load_workbook

workbook = load_workbook(filename="Factions.xlsx")
sheet = workbook.active
# Sheets
option = str(input("Enter which sheet to choose: "))
Factions = ["Mythicals", "Kingdom","Sorcerers", "Oceanics", "Alchemists", "Descendents", "Warriors", "Dragons", "Phasmas", "Titans"]

i = 0
while i < len(Factions):
    if option is Factions[i]:
        sheet = workbook[option]
        break
    else:
        i += 1

#######################################

# Categories
cardTypes = "A"
tier = "B"
_type = "C"
name = "D"
dmg = "E"
ability = "F"
timeAt = "G"

num = sheet.max_row+1

# Formulas
cardTypeForm = f'=IF(B{num}>0, \"Summon\", \"Invocation\")'
dmgForm = f'=IFS(C{num}=\"S\", B{num}*2, C{num}=\"T\", B{num}*2-1)'
timeAtForm = "=TODAY()"

# So when it comes to making a card ask whether for Summon or Invocation
CT = str(input("Card Type: "))
CT = CT.lower()
if CT is "summon":
    # When summon we need Tier int, Type str, Name str, Dmg int, and Ability str
_tier = int(input("Enter Tier: "))
__type = str(input("Enter Type: "))
_name = str(input(""))