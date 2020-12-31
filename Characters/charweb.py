from openpyxl import load_workbook

workbook = load_workbook(filename="Characters.xlsx")
sheet = workbook.active

Name = "A"
Age = "B"
Gender = "C"
Faction = "D"
Role = "E"

size = sheet.max_row

nameArr = []
ageArr = []
genderArr = []
factionArr = []
roleArr = []

i = 2
j = 1
while j < size:
    nameArr.append(sheet[Name + str(i)].value)
    ageArr.append(sheet[Age + str(i)].value)
    genderArr.append(sheet[Gender + str(i)].value)
    factionArr.append(sheet[Faction + str(i)].value)
    roleArr.append(sheet[Role + str(i)].value)
    i += 1
    j += 1

# Now to itterate and create the html page
file = open("tmp.html", "w")

file.write("<table id=\"myTable\">\n<tr class=\"header\">\n\t<th style=\"width:50%;\">Name</th>\n\t<th style=\"width: "
           "50%\">Faction</th>\n</tr>\n")
k = 0
while k < len(nameArr):
    file.write(f'<tr>\n\t<th><a href=\"#{nameArr[k]}\">{nameArr[k]}</a></th>\n\t<th>{factionArr[k]}</th>\n</tr>\n\n')
    k += 1
file.write("</table>\n")

k = 0
while k < len(nameArr):
    file.write(f'<h1 id=\"#{nameArr[k]}\">{nameArr[k]}</h1>\n<hr>\n<h2>Age: {ageArr[k]}</h2>\n<br>\n<h2>Gender: {genderArr[k]}</h2>\n<br>\n<h2>Faction: {factionArr[k]}</h2>\n<br>\n<h2>Role: {roleArr[k]}</h2>\n<br>\n<hr>\n\n')
    k += 1
file.close()
