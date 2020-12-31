from openpyxl import load_workbook

workbook = load_workbook(filename="Characters.xlsx")
sheet = workbook.active

Name = "A"
Age = "B"
Gender = "C"
Faction = "D"
Role = "E"

# First Step is to ask for the details in order.
charName = str(input("Character Name: "))
charAge = int(input("Character Age: "))
charGender = str(input("Character Gender: "))
charFaction = str(input("Character Faction: "))
charRole = str(input("Character Role: "))

count = sheet.max_row + 1
count = str(count)
sheet[Name + count] = charName
sheet[Age + count] = charAge
sheet[Gender + count] = charGender
sheet[Faction + count] = charFaction
sheet[Role + count] = charRole
workbook.save(filename="Characters.xlsx")
